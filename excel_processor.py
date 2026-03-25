#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
excel_processor.py
══════════════════════════════════════════════════════════════════════════════
International Building Number Extractor — Excel Processor
══════════════════════════════════════════════════════════════════════════════

PURPOSE
-------
Read an Excel file containing SAP address columns CLIADDLN3_LA and
CLIADDLN4_LA, concatenate them to form the full street address, extract the
building / house / street number from each row, and write the enriched file.

USAGE
-----
  # From command line:
  python excel_processor.py --input  addresses.xlsx \
                             --output output_with_numbers.xlsx \
                             --sheet  Sheet1           \   # optional (default: first sheet)
                             --col3   CLIADDLN3_LA     \   # optional (default name)
                             --col4   CLIADDLN4_LA     \   # optional (default name)
                             --out-col BUILDING_NUMBER     # optional (default name)

  # From Python:
  from excel_processor import process_excel
  process_excel("addresses.xlsx", "output.xlsx")

OUTPUT COLUMNS ADDED
--------------------
  BUILDING_NUMBER   → extracted number (e.g. "12", "12A", "12bis", "10-12")
  BN_CONFIDENCE     → float 0.0-1.0 indicating extraction reliability
  BN_LANG           → detected language/script code (FR, DE, IT, ES, …)
  BN_METHOD         → human-readable description of the extraction method
  FULL_ADDRESS      → concatenated CLIADDLN3_LA + CLIADDLN4_LA (for audit)

REQUIREMENTS
------------
  pip install openpyxl pandas
"""

import sys
import argparse
import logging
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Import our parser (works whether run from project root or package dir) ──
try:
    from address_parser import InternationalBuildingNumberExtractor
except ImportError:
    sys.path.insert(0, str(Path(__file__).parent))
    from address_parser import InternationalBuildingNumberExtractor


# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-8s │ %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("excel_processor")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _safe_str(val) -> str:
    """Convert cell value to string; return empty string for NaN/None."""
    if val is None:
        return ""
    if isinstance(val, float):
        import math
        if math.isnan(val):
            return ""
    return str(val).strip()


def _find_column(df: pd.DataFrame, name: str) -> str:
    """
    Case-insensitive column search.
    Raises KeyError with a helpful message if not found.
    """
    mapping = {c.lower().strip(): c for c in df.columns}
    key = name.lower().strip()
    if key in mapping:
        return mapping[key]
    raise KeyError(
        f"Column '{name}' not found.  "
        f"Available columns: {list(df.columns)}"
    )


def _confidence_fill(conf: float) -> PatternFill:
    """Return a background fill colour based on confidence level."""
    if conf >= 0.90:
        hex_color = "C6EFCE"   # green
    elif conf >= 0.70:
        hex_color = "FFEB9C"   # yellow
    elif conf > 0.0:
        hex_color = "FFC7CE"   # red/orange
    else:
        hex_color = "D9D9D9"   # grey (not found)
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _apply_header_style(ws, header_row_idx: int, new_col_indices: list):
    """Style the new header cells."""
    header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in new_col_indices:
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border


# ─────────────────────────────────────────────────────────────────────────────
# Core processing function
# ─────────────────────────────────────────────────────────────────────────────

def process_excel(
    input_path:  str,
    output_path: str  = None,
    sheet_name:  str  = None,
    col3:        str  = "CLIADDLN3_LA",
    col4:        str  = "CLIADDLN4_LA",
    out_col:     str  = "BUILDING_NUMBER",
    add_confidence: bool = True,
    add_lang:       bool = True,
    add_method:     bool = True,
    add_full_addr:  bool = True,
) -> pd.DataFrame:
    """
    Process the Excel file and return an enriched DataFrame.

    Parameters
    ----------
    input_path   : path to the input .xlsx / .xls file
    output_path  : path for the output file (if None, derived from input_path)
    sheet_name   : sheet to read (None = first sheet)
    col3         : column name for CLIADDLN3_LA  (can be customized)
    col4         : column name for CLIADDLN4_LA  (can be customized)
    out_col      : name for the building number output column
    add_confidence / add_lang / add_method / add_full_addr : toggle extra cols
    """

    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if output_path is None:
        output_path = input_path.parent / (input_path.stem + "_with_numbers.xlsx")
    output_path = Path(output_path)

    # ── Read Excel ────────────────────────────────────────────────────────────
    log.info(f"Reading: {input_path}")
    read_kwargs = {}
    if sheet_name:
        read_kwargs["sheet_name"] = sheet_name

    df = pd.read_excel(input_path, dtype=str, **read_kwargs)
    log.info(f"Loaded {len(df):,} rows × {len(df.columns)} columns")

    # ── Locate source columns ─────────────────────────────────────────────────
    actual_col3 = _find_column(df, col3)
    actual_col4 = _find_column(df, col4)
    log.info(f"Source columns: '{actual_col3}' + '{actual_col4}'")

    # ── Run extraction ────────────────────────────────────────────────────────
    extractor = InternationalBuildingNumberExtractor()

    results = []
    for idx, row in df.iterrows():
        ln3 = _safe_str(row[actual_col3])
        ln4 = _safe_str(row[actual_col4])
        res = extractor.extract(ln3, ln4)
        results.append(res)

        if (idx + 1) % 500 == 0:
            log.info(f"  Processed {idx + 1:,} / {len(df):,} rows …")

    log.info(f"Extraction complete ({len(results):,} rows)")

    # ── Build output columns ──────────────────────────────────────────────────
    df[out_col] = [r.building_number if r.building_number else "" for r in results]

    if add_full_addr:
        df["FULL_ADDRESS"]  = [r.raw_address for r in results]
    if add_confidence:
        df["BN_CONFIDENCE"] = [round(r.confidence, 2) for r in results]
    if add_lang:
        df["BN_LANG"]       = [r.lang_code or "" for r in results]
    if add_method:
        df["BN_METHOD"]     = [r.method for r in results]

    # ── Statistics ────────────────────────────────────────────────────────────
    found        = sum(1 for r in results if r.building_number)
    high_conf    = sum(1 for r in results if r.confidence >= 0.90)
    med_conf     = sum(1 for r in results if 0.70 <= r.confidence < 0.90)
    low_conf     = sum(1 for r in results if 0.0 < r.confidence < 0.70)
    not_found    = sum(1 for r in results if not r.building_number)

    log.info("─" * 55)
    log.info(f"  Total rows         : {len(df):>8,}")
    log.info(f"  Numbers found      : {found:>8,}  ({100*found/len(df):.1f}%)")
    log.info(f"  High confidence    : {high_conf:>8,}  (≥ 0.90)")
    log.info(f"  Medium confidence  : {med_conf:>8,}  (0.70–0.89)")
    log.info(f"  Low confidence     : {low_conf:>8,}  (< 0.70)")
    log.info(f"  Not found          : {not_found:>8,}")
    log.info("─" * 55)

    # ── Write to Excel with styling ───────────────────────────────────────────
    log.info(f"Writing: {output_path}")

    df.to_excel(output_path, index=False, engine="openpyxl")

    # Post-process: apply conditional formatting on BN_CONFIDENCE column
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    header = {cell.value: cell.column for cell in ws[1]}

    # Identify new columns
    new_cols = [out_col]
    if add_full_addr:  new_cols.append("FULL_ADDRESS")
    if add_confidence: new_cols.append("BN_CONFIDENCE")
    if add_lang:       new_cols.append("BN_LANG")
    if add_method:     new_cols.append("BN_METHOD")

    new_col_indices = [header[c] for c in new_cols if c in header]
    _apply_header_style(ws, 1, new_col_indices)

    # Colour-code confidence column
    if add_confidence and "BN_CONFIDENCE" in header:
        conf_col = header["BN_CONFIDENCE"]
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=conf_col)
            try:
                conf_val = float(cell.value) if cell.value != "" else 0.0
            except (ValueError, TypeError):
                conf_val = 0.0
            cell.fill = _confidence_fill(conf_val)

    # Colour-code BUILDING_NUMBER column: highlight empty cells
    if out_col in header:
        bn_col = header[out_col]
        empty_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
        found_fill = PatternFill(start_color="EBF5EB", end_color="EBF5EB", fill_type="solid")
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=bn_col)
            cell.fill = found_fill if cell.value else empty_fill
            cell.font = Font(bold=True, name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths (approximate)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info(f"Done → {output_path}")

    return df


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────────────────────

def _parse_args():
    p = argparse.ArgumentParser(
        description="Extract building numbers from CLIADDLN3_LA + CLIADDLN4_LA columns.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--input",   "-i", required=True,           help="Input Excel file path")
    p.add_argument("--output",  "-o", default=None,            help="Output Excel file path (default: <input>_with_numbers.xlsx)")
    p.add_argument("--sheet",   "-s", default=None,            help="Sheet name (default: first sheet)")
    p.add_argument("--col3",          default="CLIADDLN3_LA",  help="Column name for address line 3")
    p.add_argument("--col4",          default="CLIADDLN4_LA",  help="Column name for address line 4")
    p.add_argument("--out-col",       default="BUILDING_NUMBER",help="Output column name for building number")
    p.add_argument("--no-confidence", action="store_true",      help="Omit BN_CONFIDENCE column")
    p.add_argument("--no-lang",       action="store_true",      help="Omit BN_LANG column")
    p.add_argument("--no-method",     action="store_true",      help="Omit BN_METHOD column")
    p.add_argument("--no-full-addr",  action="store_true",      help="Omit FULL_ADDRESS column")
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    process_excel(
        input_path      = args.input,
        output_path     = args.output,
        sheet_name      = args.sheet,
        col3            = args.col3,
        col4            = args.col4,
        out_col         = args.out_col,
        add_confidence  = not args.no_confidence,
        add_lang        = not args.no_lang,
        add_method      = not args.no_method,
        add_full_addr   = not args.no_full_addr,
    )
